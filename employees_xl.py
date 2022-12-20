from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, colors, Border, Side, Font

# columns starting from Column A
FINAL_COLUMNS = [
    'No.', 'ID', 'First Name', 'Last Name',
    'Department', 'Business Unit', 'Job Title',
    'Gender', 'Ethnicity', 'Age',
    'Country', 'City', 'Hire Date',
    'Annual Salary', 'Exit Date'
]

root = Path().cwd()
new = root / 'new_data'
new.mkdir(exist_ok=True)

df = pd.read_excel('Employee Sample.xlsx', engine='openpyxl')
# rename columns
df.rename(columns={
    'EEID': 'ID'
}, inplace=True)

FIRST_NAME = []
LAST_NAME = []

fu_name_list = df["Full Name"].to_list()
for name in fu_name_list:
    f_name = name.split(" ")[0]
    FIRST_NAME.append(f_name)
    l_name = name.split(" ")[1]
    LAST_NAME.append(l_name)

df.drop(columns=['Full Name', 'Bonus %'], inplace=True, errors='ignore')

f_name_series = pd.Series(FIRST_NAME, name='First Name')
l_name_series = pd.Series(LAST_NAME, name='Last Name')
index_no = pd.Series(range(1, len(df)+1), name='No.')

df = df.join(f_name_series)
df = df.join(l_name_series)
df = df.join(index_no)
df = df.reindex(columns=FINAL_COLUMNS)

df.to_excel('Employee Data Sample.xlsx', sheet_name='Employee Data', index=False)

wb = load_workbook('Employee Data Sample.xlsx')
sheet = wb[wb.sheetnames[0]]
sheet.freeze_panes = 'A2'

# openpyxl.styles
font_ = Font(name='Times New Roman',
             size=12,
             bold=sheet[1],
             color=colors.Color("00003366")
             )

fill_ = PatternFill(fgColor=colors.Color("00FFFFCC"),
                    patternType="solid"
                    )

redFill = PatternFill(start_color='FFAAAA', end_color='FFAAAA', patternType='solid')

border_ = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin')
                 )

alignment_ = Alignment(horizontal='general',
                       vertical='center'
                       )

for idx, cell in enumerate(sheet[1], 1):
    cell.fill = fill_
    cell.font = font_
    cell.border = border_
    cell.alignment = alignment_

for col in sheet.iter_cols():
    first_cell = col[0]
    column_letter = first_cell.column_letter
    column_name = str(first_cell.value).lower()

    if column_name in ['first name', 'last name']:
        sheet.column_dimensions[column_letter].width = 13
    elif column_name in ['no.', 'age']:
        sheet.column_dimensions[column_letter].width = 9
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    elif column_name in ['hire date', 'exit date']:
        sheet.column_dimensions[column_letter].width = 11
        for cell in col:
            cell.number_format = 'dd/mm/yyyy'
            cell.alignment = Alignment(horizontal='left', vertical='center')
    elif column_name in ['annual salary']:
        sheet.column_dimensions[column_letter].width = 13
        for cell in col:
            cell.number_format = '$#,###'
            cell.alignment = Alignment(horizontal='right', vertical='center')

wb.save(new.joinpath('Employee Data Sample.xlsx'))






